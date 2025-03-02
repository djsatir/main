import sqlite3
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import io

# Подключение к базе данных SQLite
def init_db():
    conn = sqlite3.connect('budget.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS budget (
                 id INTEGER PRIMARY KEY AUTOINCREMENT,
                 user TEXT,
                 date TEXT,
                 category TEXT,
                 amount INTEGER)''')
    conn.commit()
    conn.close()

# Сохранение записи в базу данных
def save_to_db(user, date, category, amount):
    conn = sqlite3.connect('budget.db')
    c = conn.cursor()
    c.execute("INSERT INTO budget (user, date, category, amount) VALUES (?, ?, ?, ?)",
              (user, date, category, amount))
    conn.commit()
    conn.close()

# Извлечение числа из сообщения с валидацией
def extract_amount(text):
    match = re.search(r'^[+-]\d+$', text.strip())  # Только +число или -число
    if match:
        return int(match.group())
    return None

# Обработка текстовых сообщений
def handle_message(update, context):
    user = update.message.from_user.username or str(update.message.from_user.id)
    text = update.message.text
    date = update.message.date.strftime('%Y-%m-%d')

    amount = extract_amount(text)
    if amount is not None:
        category = 'Доход' if amount > 0 else 'Расход'
        save_to_db(user, date, category, amount)
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text=f"Записано: {user}, {category}: {amount} ₽")
    else:
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="Неверный формат. Используйте: +500 или -700")

# Получение статистики за день
def stats_day(update, context):
    date = datetime.now().strftime('%Y-%m-%d')
    conn = sqlite3.connect('budget.db')
    c = conn.cursor()
    c.execute("SELECT user, category, SUM(amount) FROM budget WHERE date = ? GROUP BY user, category", (date,))
    rows = c.fetchall()
    conn.close()

    stats = {}
    for user, category, total in rows:
        if user not in stats:
            stats[user] = {'Доход': 0, 'Расход': 0}
        stats[user][category] = total

    response = "Статистика за сегодня:\n"
    for user, data in stats.items():
        income = data['Доход']
        expense = data['Расход']
        balance = income + expense
        response += f"{user}:\n- Доход: {income} ₽\n- Расход: {expense} ₽\n- Баланс: {balance} ₽\n\n"
    
    if not stats:
        response = "Сегодня нет записей."
    context.bot.send_message(chat_id=update.effective_chat.id, text=response)

# Получение статистики за неделю
def stats_week(update, context):
    conn = sqlite3.connect('budget.db')
    c = conn.cursor()
    c.execute("SELECT user, category, SUM(amount) FROM budget WHERE date >= date('now', '-7 days') GROUP BY user, category")
    rows = c.fetchall()
    conn.close()

    stats = {}
    for user, category, total in rows:
        if user not in stats:
            stats[user] = {'Доход': 0, 'Расход': 0}
        stats[user][category] = total

    response = "Статистика за неделю:\n"
    for user, data in stats.items():
        income = data['Доход']
        expense = data['Расход']
        balance = income + expense
        response += f"{user}:\n- Доход: {income} ₽\n- Расход: {expense} ₽\n- Баланс: {balance} ₽\n\n"
    
    if not stats:
        response = "За неделю нет записей."
    context.bot.send_message(chat_id=update.effective_chat.id, text=response)

# Получение статистики за произвольный период
def stats_period(update, context):
    try:
        args = context.args
        if len(args) != 2:
            raise ValueError("Укажите две даты: /stats_period YYYY-MM-DD YYYY-MM-DD")
        
        start_date = args[0]
        end_date = args[1]
        
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')

        conn = sqlite3.connect('budget.db')
        c = conn.cursor()
        c.execute("SELECT user, category, SUM(amount) FROM budget WHERE date BETWEEN ? AND ? GROUP BY user, category",
                  (start_date, end_date))
        rows = c.fetchall()
        conn.close()

        stats = {}
        for user, category, total in rows:
            if user not in stats:
                stats[user] = {'Доход': 0, 'Расход': 0}
            stats[user][category] = total

        response = f"Статистика за период {start_date} - {end_date}:\n"
        for user, data in stats.items():
            income = data['Доход']
            expense = data['Расход']
            balance = income + expense
            response += f"{user}:\n- Доход: {income} ₽\n- Расход: {expense} ₽\n- Баланс: {balance} ₽\n\n"
        
        if not stats:
            response = "За указанный период нет записей."
    except ValueError as e:
        response = str(e) if "Укажите две даты" in str(e) else "Ошибка формата дат. Используйте: /stats_period YYYY-MM-DD YYYY-MM-DD"
    
    context.bot.send_message(chat_id=update.effective_chat.id, text=response)

# Экспорт данных в Excel
def export(update, context):
    conn = sqlite3.connect('budget.db')
    c = conn.cursor()
    c.execute("SELECT * FROM budget")
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    headers = ["ID", "Пользователь", "Дата", "Категория", "Сумма (₽)"]
    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    for row_num, row in enumerate(rows, 2):
        for col_num, value in enumerate(row, 1):
            ws[f"{get_column_letter(col_num)}{row_num}"] = value

    filename = "budget_export.xlsx"
    wb.save(filename)
    with open(filename, 'rb') as file:
        context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename=filename)

    context.bot.send_message(chat_id=update.effective_chat.id, text="Данные экспортированы в budget_export.xlsx")

# Генерация графика за произвольный период с фильтром по пользователю
def plot_period(update, context):
    try:
        args = context.args
        if len(args) not in [2, 3]:
            raise ValueError("Используйте: /plot_period [username] YYYY-MM-DD YYYY-MM-DD")

        if len(args) == 3:
            username = args[0]
            start_date = args[1]
            end_date = args[2]
        else:
            username = None
            start_date = args[0]
            end_date = args[1]

        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')

        conn = sqlite3.connect('budget.db')
        c = conn.cursor()
        if username:
            c.execute("SELECT user, category, SUM(amount) FROM budget WHERE user = ? AND date BETWEEN ? AND ? GROUP BY user, category",
                      (username, start_date, end_date))
        else:
            c.execute("SELECT user, category, SUM(amount) FROM budget WHERE date BETWEEN ? AND ? GROUP BY user, category",
                      (start_date, end_date))
        rows = c.fetchall()
        conn.close()

        stats = {}
        for user, category, total in rows:
            if user not in stats:
                stats[user] = {'Доход': 0, 'Расход': 0}
            stats[user][category] = total

        if not stats:
            context.bot.send_message(chat_id=update.effective_chat.id, text="За указанный период нет данных для графика.")
            return

        # Подготовка данных для графика
        users = list(stats.keys())
        incomes = [stats[user]['Доход'] for user in users]
        expenses = [abs(stats[user]['Расход']) for user in users]  # Расходы в положительных значениях для графика

        # Создание графика
        fig, ax = plt.subplots(figsize=(8, 6))
        bar_width = 0.35
        x = range(len(users))

        plt.bar(x, incomes, bar_width, label='Доход', color='green')
        plt.bar([i + bar_width for i in x], expenses, bar_width, label='Расход', color='red')

        plt.xlabel('Пользователи')
        plt.ylabel('Сумма (₽)')
        title = f'Доходы и расходы за {start_date} - {end_date}'
        if username:
            title += f' (только {username})'
        plt.title(title)
        plt.xticks([i + bar_width / 2 for i in x], users)
        plt.legend()

        # Сохранение графика в буфер
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()

        # Отправка графика в чат
        context.bot.send_photo(chat_id=update.effective_chat.id, photo=buf, filename='budget_plot.png')
        buf.close()

    except ValueError as e:
        response = str(e) if "Используйте" in str(e) else "Ошибка формата дат. Используйте: /plot_period [username] YYYY-MM-DD YYYY-MM-DD"
        context.bot.send_message(chat_id=update.effective_chat.id, text=response)

# Помощь по командам
def help_command(update, context):
    response = "Команды бота:\n" \
               "/stats_day - Статистика за сегодня\n" \
               "/stats_week - Статистика за неделю\n" \
               "/stats_period YYYY-MM-DD YYYY-MM-DD - Статистика за период\n" \
               "/plot_period [username] YYYY-MM-DD YYYY-MM-DD - График за период (опционально для пользователя)\n" \
               "/export - Экспорт данных в Excel\n" \
               "Просто отправьте сообщение с числом, например: +500 или -700"
    context.bot.send_message(chat_id=update.effective_chat.id, text=response)

# Основная функция
def main():
    # Инициализация базы данных
    init_db()

    # Замените 'YOUR_TOKEN' на токен вашего бота
    updater = Updater("7372121610:AAEmbJ-99sYuOKAWB4S1cvjUPNn5z3mr32o", use_context=True)

    # Регистрация обработчиков
    dp = updater.dispatcher
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command, handle_message))
    dp.add_handler(CommandHandler("stats_day", stats_day))
    dp.add_handler(CommandHandler("stats_week", stats_week))
    dp.add_handler(CommandHandler("stats_period", stats_period))
    dp.add_handler(CommandHandler("plot_period", plot_period))
    dp.add_handler(CommandHandler("export", export))
    dp.add_handler(CommandHandler("help", help_command))

    # Запуск бота
    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()
